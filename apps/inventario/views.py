from django.shortcuts import render
from django.views.generic import CreateView , ListView,UpdateView, DeleteView,TemplateView
from django.views.generic.detail import DetailView
from apps.inventario.models import usuario ,equipo , asignacion , pda , telefono
from apps.inventario.forms import UsuarioForm , EquipoForm , asignacionForm, pdaForm, telefonoForm
from django.urls import reverse_lazy
import pyrebase
from django.contrib import auth
from .filters import equipoFilter, asignacionFilter ,pdaFilter
from openpyxl import Workbook
from django.http.response import HttpResponse
# Create your views here.
#vista de pagina home y login **************************************************************************************
def home(request):
    return render(request, 'home/homeLogin.html')

def informe(request):
    return render(request,'informe/informe.html')


#vistas de usuario *************************************************************************************************
class UsuarioCreate(CreateView):
    model = usuario
    form_class = UsuarioForm
    template_name = 'usuario/Usuario_form.html'
    success_url = reverse_lazy('usuario:usuario_listar')
 
class UsuarioList(ListView):
    model = usuario
    template_name = 'usuario/usuario_list.html'
    paginate_by = 10

class UsuarioUpdate(UpdateView):
    model = usuario
    form_class = UsuarioForm 
    template_name = 'usuario/Usuario_form.html'
    success_url = reverse_lazy('usuario:usuario_listar') 

class usuarioDelete(DeleteView):
    model = usuario
    template_name = 'usuario/usuario_delete.html'
    success_url = reverse_lazy('usuario:usuario_listar')

class usuarioShow(DetailView):
    model = usuario
    template_name = 'usuario/usuario_show.html'



#listas para agregar y editar equipamiento **********************************************************************
class EquipamientoCreate(CreateView):
    model = equipo
    form_class = EquipoForm
    template_name = 'equipos/equipo_form.html'
    success_url = reverse_lazy('equipos:equipo_listar')

class equipoList(ListView):
    model = equipo
    template_name = 'equipos/equipo_list.html'
    paginate_by = 20
   

class equipoUpdate(UpdateView):
    model = equipo
    form_class = EquipoForm
    template_name = 'equipos/equipo_form.html'
    success_url = reverse_lazy('equipos:equipo_listar')

class equipoDelete(DeleteView):
    model = equipo
    template_name = 'equipos/equipo_delete.html'
    success_url = reverse_lazy('equipos:equipo_listar')

class equipoShow(DetailView):
    model = equipo
    template_name = 'equipos/equipo_show.html'

def equipoBuscar(request):
    f = equipoFilter(request.GET, queryset= equipo.objects.all())
    return render(request , 'equipos/equipo_search.html' , {'filter' : f})
  
class ReporteEquipoExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        _equipo = equipo.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE EQUIPOS'

        ws.merge_cells('B1:E1')
        ws['B3'] = 'Marca'
        ws['C3'] = 'Modelo'
        ws['D3'] = 'Serie'
        ws['E3'] = 'Rotulo'
        ws['F3'] = 'Fecha ingreso'
        ws['G3'] = 'Tipo'
        ws['H3'] = 'Descripcion'

        cont = 4
        for equipos in _equipo:
            ws.cell(row = cont, column = 2).value = equipos.marca
            ws.cell(row = cont, column = 3).value = equipos.modelo
            ws.cell(row = cont, column = 4).value = equipos.serie
            ws.cell(row = cont, column = 5).value = equipos.rotulo
            ws.cell(row = cont, column = 6).value = equipos.fecha_ing
            ws.cell(row = cont, column = 7).value = equipos.tipo
            ws.cell(row = cont, column = 8).value = equipos.descripcion
            cont+=1

        nombre_archivo = "ReporteEquipoExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

#listas para agregar y editar asignacion de equipamiento ******************************************************************
class AsignacionCreate(CreateView):
    model = asignacion
    form_class = asignacionForm
    template_name = 'asignacion/asignacion_form.html'
    success_url = reverse_lazy('Asignacion:asignacion_listar')

class asignacionList(ListView):
    model = asignacion
    template_name = 'asignacion/asignacion_list.html'
    paginate_by = 15

class asignacionUpdate(UpdateView):
    model = asignacion
    form_class = asignacionForm
    template_name = 'asignacion/asignacion_form.html'
    success_url = reverse_lazy('Asignacion:asignacion_listar')

class asignacionDelete(DeleteView):
    model = asignacion
    template_name = 'asignacion/asignacion_delete.html'
    success_url = reverse_lazy('Asignacion:asignacion_listar')

class asignacionShow(DetailView):
    model = asignacion
    template_name = 'asignacion/asignacion_show.html'

def asignacionBuscar(request):
    asig = asignacionFilter(request.GET, queryset= asignacion.objects.all())
    return render(request , 'asignacion/asignacion_search.html' , {'filter' : asig})
    

class ReporteAsignacionExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        _asig = asignacion.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE ASIGNACIONES'

        ws.merge_cells('B1:E1')
        ws['B3'] = 'Usuario asignado'
        ws['C3'] = 'Cargo'
        ws['D3'] = 'Departamento'
        ws['E3'] = 'Fecha_entrega'
        ws['F3'] = 'Marca_equipo_asignado'
        ws['G3'] = 'Modelo_equipo_asignado'
        ws['H3'] = 'Serie_equipo_asignado'
        ws['I3'] = 'Rotulo_equipo_asignado'
        ws['J3'] = 'Tipo_equipo'
        ws['K3'] = 'Tipo_equipo'

        cont = 4
        for asig in _asig:
            ws.cell(row = cont, column = 2).value = asig.usuario_asig
            ws.cell(row = cont, column = 3).value = asig.cargo
            ws.cell(row = cont, column = 4).value = asig.departamento
            ws.cell(row = cont, column = 5).value = asig.fecha_entrega
            ws.cell(row = cont, column = 6).value = asig.marca_equipo_asig
            ws.cell(row = cont, column = 7).value = asig.modelo_equipo_asig
            ws.cell(row = cont, column = 8).value = asig.serie_equipo_asig
            ws.cell(row = cont, column = 9).value = asig.rotulo_equipo_asig
            ws.cell(row = cont, column = 10).value = asig.tipo_equipo
            ws.cell(row = cont, column = 11).value = asig.activo

        nombre_archivo = "ReporteAsignacionExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response        


#listas para agregar y editar pda ******************************************************************************
class pdaCreate(CreateView):
    model = pda
    form_class = pdaForm
    template_name = 'pda/pda_form.html'
    success_url = reverse_lazy('pda:pda_listar')

class pdaList(ListView):
    model = pda
    template_name = 'pda/pda_list.html'
    paginate_by = 20

class pdaUpdate(UpdateView):
    model = pda
    form_class = pdaForm
    template_name = 'pda/pda_form.html'
    success_url = reverse_lazy('pda:pda_listar')

class pdaDelete(DeleteView):
    model = pda
    template_name = 'pda/pda_delete.html'
    success_url = reverse_lazy('pda:pda_listar')

class pdaShow(DetailView):
    model = pda
    template_name = 'pda/pda_show.html'

def pdaBuscar(request):
    p = pdaFilter(request.GET, queryset= pda.objects.all())
    return render(request , 'pda/pda_search.html' , {'filter' : p})

class ReportePdaExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        _pda = pda.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE PDA'

        ws.merge_cells('B1:E1')
        ws['B3'] = 'OT'
        ws['C3'] = 'modelo_pda'
        ws['D3'] = 'serie_pda'
        ws['E3'] = 'fecha_ing'
        ws['F3'] = 'fecha_desp'
        ws['G3'] = 'serie_bateria'
        ws['H3'] = 'num_inventario'
        ws['I3'] = 'sucursal'
        ws['J3'] = 'region'
        ws['K3'] = 'estado'
        ws['L3'] = 'dias_st'
        ws['M3'] = 'descripcion_pda'

        cont = 4
        for PDAS in _pda:
            ws.cell(row = cont, column = 2).value = PDAS.ot
            ws.cell(row = cont, column = 3).value = PDAS.modelo_pda
            ws.cell(row = cont, column = 4).value = PDAS.serie_pda
            ws.cell(row = cont, column = 5).value = PDAS.fecha_ing
            ws.cell(row = cont, column = 6).value = PDAS.fecha_desp
            ws.cell(row = cont, column = 7).value = PDAS.serie_bateria
            ws.cell(row = cont, column = 8).value = PDAS.num_inventario
            ws.cell(row = cont, column = 9).value = PDAS.sucursal
            ws.cell(row = cont, column = 10).value = PDAS.region
            ws.cell(row = cont, column = 11).value = PDAS.estado
            ws.cell(row = cont, column = 12).value = PDAS.dias_st
            ws.cell(row = cont, column = 13).value = PDAS.descripcion_pda

        nombre_archivo = "ReportePDAExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

#listas para agregar y editar teléfono **************************************************************
class telefonoCreate(CreateView):
    model = telefono
    form_class = telefonoForm
    template_name = 'telefono/telefono_form.html'
    success_url = reverse_lazy('telefono:telefono_listar')

class telegonoList(ListView):
    model = telefono
    template_name = 'telefono/telefono_list.html'
    paginate_by = 10

class telefonoUpdate(UpdateView):
    model = telefono
    form_class = telefonoForm
    template_name = 'telefono/telefono_form.html'
    success_url = reverse_lazy('telefono:telefono_listar')

class telefonoDelete(DeleteView):
    model = telefono
    template_name = 'telefono/telefono_delete.html'
    success_url = reverse_lazy('telefono:telefono_listar')

class telefonoShow(DetailView):
    model = telefono
    template_name = 'telefono/telefono_show.html'

class ReporteTelefonoExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        _telef = telefono.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE TELEFONOS'

        ws.merge_cells('B1:E1')
        ws['B3'] = 'marca_tel'
        ws['C3'] = 'modelo_tel'
        ws['D3'] = 'serie_tel'
        ws['E3'] = 'numero_tel'
        ws['F3'] = 'fecha_ing_tel'
        ws['G3'] = 'usuario_asig_tel'
        ws['H3'] = 'departamento_usuario'
      

        cont = 4
        for phone in _telef:
            ws.cell(row = cont, column = 2).value = phone.marca_tel
            ws.cell(row = cont, column = 3).value = phone.modelo_tel
            ws.cell(row = cont, column = 4).value = phone.serie_tel
            ws.cell(row = cont, column = 5).value = phone.numero_tel
            ws.cell(row = cont, column = 6).value = phone.fecha_ing_tel
            ws.cell(row = cont, column = 7).value = phone.usuario_asig_tel
            ws.cell(row = cont, column = 8).value = phone.departamento_usuario
            

        nombre_archivo = "ReporteTelefonoExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

#codigo firebase para login de usuario ************************************************************************
config = {
    
    'apiKey': "AIzaSyDmQIep3GlGhWWnllaGxIztt8NjROcGMAU",
    'authDomain': "inventario-a66cb.firebaseapp.com",
    'databaseURL': "https://inventario-a66cb.firebaseio.com",
    'projectId': "inventario-a66cb",
    'storageBucket': "inventario-a66cb.appspot.com",
    'messagingSenderId': "721216451889"

}
firebase = pyrebase.initialize_app(config)
authe = firebase.auth()

def singIn(request):

    return render(request, 'home/homeLogin.html')

def postsign(request):
    email = request.POST.get('email')
    passw = request.POST.get("password")
    try:
        user = authe.sign_in_with_email_and_password(email,passw)
    except:
        message = "Contraseña invalida, por favor intentelo nuevamente" 
        return render(request, 'home/homeLogin.html',{"messg":message})

    print(user['idToken'])
    session_id = user['idToken']
    request.session['uid']=str(session_id)
    return render(request, 'home/home.html', {"e":email})

def logout(request):
    auth.logout(request)
    return render(request,'home/homeLogin.html')
